"save SQL result-set in Excel format"
import datetime as dt
from argparse import ArgumentParser
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable, Optional, TypeVar

from openpyxl import Workbook
from openpyxl.cell import Cell, WriteOnlyCell  # type: ignore
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter  # type: ignore

T = TypeVar("T")


def export(rows: Iterable[tuple[Any, ...]], headers: list[str], types: list[type], output: Path) -> None:
    "export data in MS Excel (xlsx) format to file"

    def num_fmt(t: type) -> str:
        "Return numbering format from type"
        if issubclass(t, int):
            return "#,##0"
        if issubclass(t, (float, Decimal)):
            return "#,##0.00"
        if issubclass(t, dt.datetime):
            return "yyyy-mm-dd hh:mm:ss"
        if issubclass(t, dt.date):
            return "yyyy-mm-dd"
        if issubclass(t, dt.time):
            return "hh:mm:ss"
        return ""

    col_numfmt = [num_fmt(t) for t in types]

    wb = Workbook(write_only=True)
    wb.iso_dates = True
    ws = wb.create_sheet()  # type: ignore

    hdr_font, data_font = Font("Hack", bold=True), Font("Hack")

    def hdr_cell(val: str) -> Cell:
        "Formatted cell containing column title"
        cell = WriteOnlyCell(ws, value=val)
        cell.font = hdr_font
        return cell

    def data_cell(val: Optional[Any], num_fmt: Optional[str]) -> Optional[Cell]:
        "Formatted cell containing data value"
        if val is None:
            return None
        cell = WriteOnlyCell(ws, value=val)
        cell.font = data_font
        if num_fmt is not None:
            cell.number_format = num_fmt
        return cell

    ws.freeze_panes = "A2"  # type: ignore
    ws.auto_filter.ref = "A1:{}1".format(get_column_letter(len(headers)))  # type: ignore

    def unset_tz(x: T) -> T:
        "Removes timezone info from instances of datetime or time, since Excel doesn't support them"
        if isinstance(x, dt.datetime):
            return x if x.tzinfo is None else x.replace(tzinfo=None)
        if isinstance(x, dt.time):
            return x if x.tzinfo is None else x.replace(tzinfo=None)
        return x

    ws.append([hdr_cell(v) for v in headers])  # type: ignore
    rows = (tuple(unset_tz(c) for c in row) for row in rows)

    for row in rows:
        ws.append([data_cell(v, f) for v, f in zip(row, col_numfmt)])  # type: ignore

    with output.open("wb") as f:
        wb.save(f)  # type: ignore


def add_args(parser: ArgumentParser) -> Any:
    """parse user supplied args"""
    parser.add_argument("output", metavar="FILE", type=Path, help="output file name")
